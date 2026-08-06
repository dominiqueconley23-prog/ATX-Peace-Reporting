"""
ATX Peace Report — Core Builder Module
Called by both the CLI script and the Streamlit web app.

Entry point: build_report(dfs, cfg=None) -> dict
  dfs  — dict of DataFrames: participants, goals, outreach, circles,
           incidents, pre, post, followup, attestation
  cfg  — optional config overrides (see DEFAULT_CFG below)
Returns: {'bytes': excel_bytes, 'stats': {...}, 'fname': str}
"""
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Locked YTD data — do NOT recalculate prior quarters ─────────────────────
# These are the exact numbers submitted to the City. When a quarter closes,
# add its data here and bump Q_START forward.
YTD_Q1 = {
    'label': 'Q1', 'dates': 'Oct 1 – Dec 31, 2025',
    'total': 42, 'atx_peace': 10, 'tyj': 32,
    'pids_atx': {'3896','3948','4357','4412','4418','4427','4426','4419','4428','4435'},
    'pids_tyj': {'3851','3914','3909','3296','3902','3929','3930','3905','3911','3920',
                 '3908','3925','3923','3910','3903','3924','3907','3915','4437','3928',
                 '3927','3918','3922','3906','4439','4438','3913','3941','3912','4440',
                 '3926','3931'},
    '5b_pct': 0.881, 'outreach_sessions': None, 'people_reached': None,
    'events': 10, 'event_attendees': 1355, 'incidents': 2,
}
YTD_Q2 = {
    'label': 'Q2', 'dates': 'Jan 1 – Mar 31, 2026',
    'total': 26, 'atx_peace': 8, 'tyj': 18,
    'pids_atx': {'4521','4515','4520','4528','4574','4550','4532','4529'},
    'pids_tyj': {'3348','3916','4441','4443','4479','4480','4481','4483',
                 '4495','4496','4497','4498','4500','4502','4507','4509','4510','4545'},
    '5b_pct': 0.808, 'outreach_sessions': 209, 'people_reached': 270,
    'events': 14, 'event_attendees': 1976, 'incidents': 12,
}
PRIOR_PIDS = (YTD_Q1['pids_atx'] | YTD_Q1['pids_tyj'] |
              YTD_Q2['pids_atx'] | YTD_Q2['pids_tyj'])

# ── Default config — update each quarter ─────────────────────────────────────
DEFAULT_CFG = {
    'QUARTER':           'Q3',
    'FISCAL_YEAR':       'FY 2025-2026',
    'PERIOD':            'April 1 – June 30, 2026',
    'Q_START':           pd.Timestamp('2026-04-01'),
    'Q_END':             pd.Timestamp('2026-06-30'),
    'WEEK_START':        pd.Timestamp('2026-06-01'),
    'WEEK_END':          pd.Timestamp('2026-06-07'),
    'TARGET_5B':         0.50,
    'AGE_MIN':           18,
    'AGE_MAX':           40,
    'Q_LA_TARGET':       19,
    'Q_QUOTA_OUTREACH':  38,
    'Q_QUOTA_CM':        19,
    'Q_QUOTA_TOTAL':     57,
    'ANNUAL_GOAL_CM':    75,
    'ANNUAL_GOAL_OUT':   150,
    'MONTHS':            ['Apr', 'May', 'Jun'],
    'MONTH_MAP':         {4: 'Apr', 5: 'May', 6: 'Jun'},
    'QUOTA': {
        'J. Cooper': {'Apr': 3, 'May': 3, 'Jun': 3},
        'R. Herd':   {'Apr': 3, 'May': 3, 'Jun': 3},
        'N. Dunn':   {'Apr': 0, 'May': 2, 'Jun': 3},
        'K. Young':  {'Apr': 0, 'May': 0, 'Jun': 0},
    },
    'INC_NAME_MAP': {
        'Reese Herd':      'R. Herd',
        'Michael Salazar': 'M. Salazar',
        'Nijalon Dunn':    'N. Dunn',
        'Jackie Scott':    'J. Scott',
        'Sean Oliver':     'S. Oliver',
        'Jeremias Cooper': 'J. Cooper',
        'Alvin Stewart':   'A. Stewart',
        'Kerry Young':     'K. Young',
        'Sherwynn Patton': 'S. Patton',
    },
}

# ── Style constants ───────────────────────────────────────────────────────────
NAVY   = '1F3864'; LBLUE  = 'BDD7EE'; WHT    = 'FFFFFF'; LGR    = 'F2F2F2'
DRED   = 'C00000'; RBKG   = 'FFE0E0'; ORG    = 'FFD966'; YEL    = 'FFFF00'
GRN    = 'E2EFDA'; TEAL   = 'D6F4F0'; PURPLE = 'E2C4F0'; LGRY   = 'D9D9D9'
LAVENDER = 'E8D5F5'; ROLLOVER_BG = 'EAF2FF'; INC_BG = 'FFF0E0'
DARK_GRN = '375623'; DARK_ORG = 'C55A11'

# ── Cell helpers ──────────────────────────────────────────────────────────────
def brd():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def ss(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return ''
    return str(v).strip()

def H(ws, r, c, v, bg=None, fg=None, wrap=False, sz=10, bold=True):
    bg = bg or NAVY; fg = fg or WHT
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    cl.fill = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def SEC(ws, r, c1, c2, lbl, bg=None, sz=11):
    bg = bg or NAVY
    ws.merge_cells(f'{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}')
    cl = ws.cell(row=r, column=c1, value=lbl)
    cl.font = Font(name='Arial', bold=True, size=sz, color=WHT)
    cl.fill = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 18

def TBR(ws, r1, r2, c1, c2):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cl in row: cl.border = brd()

def cell(ws, r, c, v, bg=LGR, fg='000000', bold=False, wrap=False, sz=10, align='left'):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = Font(name='Arial', size=sz, bold=bold, color=fg)
    cl.fill = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return cl

def title_row(ws, r, c1, c2, text, bg=NAVY, sz=14):
    ws.merge_cells(f'{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}')
    cl = ws.cell(row=r, column=c1, value=text)
    cl.font = Font(name='Arial', bold=True, size=sz, color=WHT)
    cl.fill = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 28

def subtitle_row(ws, r, c1, c2, text):
    ws.merge_cells(f'{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}')
    cl = ws.cell(row=r, column=c1, value=text)
    cl.font = Font(name='Arial', italic=True, size=10, color='444444')
    cl.alignment = Alignment(horizontal='center', vertical='center')

def email_name(e):
    if pd.isna(e) or str(e).strip() == '': return 'Unknown'
    e = str(e).split('\n')[0].strip()
    local = e.split('@')[0].lower()
    p = local.split('.')
    if len(p) >= 2:
        return f"{p[0][0].upper()}. {p[1].capitalize()}"
    return f"{local[0].upper()}. {local[1:].capitalize()}" if len(local) > 1 else local

def corr_color(issue):
    if 'Dup' in issue or 'Duplicate' in issue: return PURPLE
    if 'Ineligible' in issue or 'Age' in issue: return RBKG
    if 'Attestation Discrepancy' in issue: return ORG
    if 'Goal' in issue or 'Missing' in issue: return ORG
    if 'Unassigned' in issue or 'Circle' in issue: return YEL
    if 'Incident' in issue: return INC_BG
    return LBLUE

def map_race(r):
    if pd.isna(r) or str(r).strip() == '': return 'Race - Balance Not Specified'
    r = str(r)
    if 'Hispanic' in r or 'Latino' in r or 'Spanish' in r: return 'Some Other Race'
    if 'Black' in r or 'African' in r: return 'Black or African American'
    if 'American Indian' in r or 'Alaskan' in r: return 'American Indian or Alaskan Native'
    if 'Asian' in r: return 'Asian'
    if 'White' in r: return 'White'
    if 'Two or More' in r: return 'Two or More Races'
    return 'Some Other Race'

def map_age_band(a):
    try: a = int(float(a))
    except: return '—'
    if a < 18: return 'Under 18'
    if a <= 24: return '18 To 24'
    if a <= 39: return '25 To 39'
    if a <= 54: return '40 To 54'
    return '55+'

def map_eth(r):
    if pd.isna(r) or str(r).strip() == '': return 'Ethnicity - Balance Not Specified'
    if 'Hispanic' in str(r) or 'Latino' in str(r) or 'Spanish' in str(r): return 'Hispanic or Latino'
    return 'Not Hispanic or Latino'


# ═════════════════════════════════════════════════════════════════════════════
# MAIN BUILDER
# ═════════════════════════════════════════════════════════════════════════════
def build_report(dfs: dict, cfg: dict = None) -> dict:
    """
    Build the full ATX Peace Excel report.

    dfs keys: participants, goals, outreach, circles, incidents,
              pre, post, followup, attestation
    cfg:      overrides for DEFAULT_CFG (any key)

    Returns: {'bytes': excel_bytes, 'stats': {...}, 'fname': str}
    """
    # ── Merge config ──────────────────────────────────────────────
    c = {**DEFAULT_CFG}
    if cfg: c.update(cfg)

    Q_START   = c['Q_START']
    Q_END     = c['Q_END']
    QUARTER   = c['QUARTER']
    FY        = c['FISCAL_YEAR']
    PERIOD    = c['PERIOD']
    TARGET_5B = c['TARGET_5B']
    AGE_MIN   = c['AGE_MIN']
    AGE_MAX   = c['AGE_MAX']
    WEEK_START      = c['WEEK_START']
    WEEK_END        = c['WEEK_END']
    PREV_WEEK_START = WEEK_START - pd.Timedelta(days=7)
    PREV_WEEK_END   = WEEK_END   - pd.Timedelta(days=7)
    QUOTA      = c['QUOTA']
    Q_LA_TARGET = c['Q_LA_TARGET']
    Q_QUOTA_CM  = c['Q_QUOTA_CM']
    Q_QUOTA_OUT = c['Q_QUOTA_OUTREACH']
    Q_QUOTA_TOT = c['Q_QUOTA_TOTAL']
    MONTHS      = c['MONTHS']
    MONTH_MAP   = c['MONTH_MAP']
    INC_NAME_MAP = c['INC_NAME_MAP']
    ANNUAL_GOAL_CM  = c['ANNUAL_GOAL_CM']
    ANNUAL_GOAL_OUT = c['ANNUAL_GOAL_OUT']
    IMPROVEMENT_RATE_TARGET = 57 / 75
    TODAY = datetime.now().strftime('%m/%d/%Y')

    # ── Load DataFrames ───────────────────────────────────────────
    parts_raw    = dfs['participants']
    goals_raw    = dfs['goals']
    outreach_raw = dfs['outreach'].copy()
    circles_raw  = dfs['circles'].copy()
    incidents_raw = dfs['incidents'].copy()
    pre_raw      = dfs['pre']
    post_raw     = dfs['post']
    followup_raw = dfs['followup']
    attest_form  = dfs['attestation']

    # ── Participants ──────────────────────────────────────────────
    atx = parts_raw[parts_raw['Department selection'].astype(str)
                    .str.contains('ATX Peace', na=False)].copy()
    atx['pid'] = atx['Participant iD'].astype(str).str.strip().str.replace('.0','',regex=False)
    atx['age'] = pd.to_numeric(atx['Current Age'], errors='coerce')
    atx['case_start'] = pd.to_datetime(atx['Date when case started'], errors='coerce')
    atx['coordinator'] = atx['Assigned Coordinator'].apply(
        lambda x: email_name(str(x).split('\n')[0] if '\n' in str(x) else x))
    atx['dual_coord'] = atx['Assigned Coordinator'].astype(str).str.contains('\n', na=False)
    atx_dedup = atx.sort_values('case_start').drop_duplicates(subset='pid').copy()

    q3_all   = atx_dedup[(atx_dedup['case_start'] >= Q_START) & (atx_dedup['case_start'] <= Q_END)].copy()
    q3_part  = q3_all[q3_all['Participant or Non-Participant'] == 'Participant'].copy()
    q3_elig  = q3_part[(q3_part['age'] >= AGE_MIN) & (q3_part['age'] <= AGE_MAX)].copy()
    q3_inelig= q3_part[(q3_part['age'] < AGE_MIN) | (q3_part['age'] > AGE_MAX)].copy()
    q3_nonpart = q3_all[q3_all['Participant or Non-Participant'] != 'Participant'].copy()
    rollover = atx_dedup[
        (atx_dedup['case_start'] < Q_START) &
        (atx_dedup['Participant or Non-Participant'] == 'Participant') &
        (atx_dedup['Case Manage Progress'] == 'In Progress')
    ].copy()

    q3_pids = set(q3_elig['pid'])
    all_coords = sorted(set(q3_elig['coordinator'].tolist() + rollover['coordinator'].tolist()))

    # ── Cross-quarter duplicates ──────────────────────────────────
    duplicate_q3 = q3_elig[q3_elig['pid'].isin(PRIOR_PIDS)].copy()

    # ── Goals ─────────────────────────────────────────────────────
    goals = goals_raw[goals_raw['Department selection'].astype(str)
                      .str.contains('ATX Peace', na=False)].copy()
    goals['pid'] = goals['Participant iD'].astype(str).str.strip().str.replace('.0','',regex=False)

    def goals_summary(pid_set):
        g = goals[goals['pid'].isin(pid_set)]
        return g.groupby('pid').agg(
            total_goals=('Goal Status','count'),
            completed=('Goal Status', lambda x: (x=='Complete').sum()),
            in_progress=('Goal Status', lambda x: (x=='In Progress').sum()),
            categories=('Goal Category', lambda x: '; '.join(sorted(x.dropna().unique()))),
            branches=('Selected Goal Branch', lambda x: '; '.join(sorted(x.dropna().unique()))),
        ).reset_index()

    q3_missing_age = q3_part[q3_part['age'].isna()].copy()
    q3_all_inelig  = pd.concat([q3_inelig, q3_missing_age], ignore_index=True).drop_duplicates(subset='pid')

    gs_q3       = goals_summary(q3_pids)
    gs_rollover = goals_summary(set(rollover['pid']))
    q3_elig   = q3_elig.merge(gs_q3, on='pid', how='left')
    rollover  = rollover.merge(gs_rollover, on='pid', how='left')
    for df in [q3_elig, rollover]:
        df['circles_attended'] = 0
        for col in ['total_goals','completed','in_progress']:
            df[col] = df[col].fillna(0).astype(int)
        for col in ['categories','branches']:
            df[col] = df[col].fillna('').astype(str)

    # ── 5B ────────────────────────────────────────────────────────
    completed_pids_q3 = set(goals[goals['Goal Status']=='Complete']['pid']) & q3_pids
    ls_num = len(completed_pids_q3)
    ls_pct = ls_num / len(q3_elig) if len(q3_elig) > 0 else 0

    # ── Follow-ups ────────────────────────────────────────────────
    fu = followup_raw[followup_raw['Department selection'].astype(str)
                      .str.contains('ATX Peace', na=False)].copy()
    fu['pid']  = fu['Participant iD'].astype(str).str.strip().str.replace('.0','',regex=False)
    fu['date'] = pd.to_datetime(fu['Date when follow up was made?'], errors='coerce')
    fu_q3 = fu[(fu['date'] >= Q_START) & (fu['date'] <= Q_END)].copy()

    # ── Assessments ───────────────────────────────────────────────
    pre  = pre_raw[pre_raw['Department selection'].astype(str).str.contains('ATX Peace', na=False)].copy()
    post = post_raw[post_raw['Department selection'].astype(str).str.contains('ATX Peace', na=False)].copy()
    pre['pid']  = pre['Participant iD'].astype(str).str.strip().str.replace('.0','',regex=False)
    post['pid'] = post['Participant iD'].astype(str).str.strip().str.replace('.0','',regex=False)
    pre_pids_all  = set(pre['pid'])
    post_pids_all = set(post['pid'])

    # ── Self-attestation ──────────────────────────────────────────
    attest_clean = attest_form[~attest_form['Full Name'].str.lower().str.contains('test', na=False)].copy()
    attest_clean['submit_date'] = pd.to_datetime(attest_clean['Date of Form Submission'], errors='coerce')
    attest_name_map = {}
    for _, r in attest_clean.iterrows():
        key = r['Full Name'].strip().lower()
        attest_name_map[key] = r['submit_date'].strftime('%m/%d/%Y') if pd.notna(r['submit_date']) else '✓'

    attest_pid_map = {}
    for _, r in atx_dedup.iterrows():
        val  = ss(r.get('Did the participant complete the self attestaion form?',''))
        date = ss(r.get('Date of Self-Attestation Completion',''))
        if val.lower() == 'yes':
            attest_pid_map[r['pid']] = date if date else '✓'

    def attest_display(pid):
        row = atx_dedup[atx_dedup['pid'] == pid]
        if not row.empty:
            name = (ss(row.iloc[0].get('First Name','')) + ' ' + ss(row.iloc[0].get('Last Name',''))).strip().lower()
            for key, date in attest_name_map.items():
                if name.split()[0] in key if name.split() else False:
                    return date
        if pid in attest_pid_map:
            return attest_pid_map[pid]
        return '⚠ Missing'

    # ── Outreach ──────────────────────────────────────────────────
    outreach_raw['date'] = pd.to_datetime(outreach_raw['Date of Outreach or Canvass Happened'], errors='coerce')
    outreach_raw['coordinator'] = outreach_raw['Created by'].apply(email_name)
    out_q3 = outreach_raw[(outreach_raw['date'] >= Q_START) & (outreach_raw['date'] <= Q_END)].copy()
    out_q3['is_canvass']  = out_q3['What did you do? Canvass or Outreach?[Canvass]'].eq(1)
    out_q3['is_outreach'] = out_q3['What did you do? Canvass or Outreach?[Outreach]'].eq(1)
    out_q3['people'] = pd.to_numeric(
        out_q3['Estimated Number of People Reached.1'].fillna(out_q3['Estimated Number of People Reached']),
        errors='coerce').fillna(0)
    out_q3['out_hrs'] = pd.to_numeric(out_q3['How long was outreach? (By the hour)'], errors='coerce').fillna(0)
    out_q3['can_hrs'] = pd.to_numeric(out_q3['How long was canvassing? (By the hour)'], errors='coerce').fillna(0)
    out_q3['notes'] = (out_q3.get('Any notes about the Outreach?', pd.Series(dtype=str)).fillna('') +
                       out_q3.get('Any notes about Canvassing?', pd.Series(dtype=str)).fillna(''))
    out_q3['out_type'] = out_q3.apply(
        lambda r: 'Both' if r['is_canvass'] and r['is_outreach'] else
                  ('Canvass' if r['is_canvass'] else 'Outreach'), axis=1)

    def out_location(row):
        locs = []
        for col in outreach_raw.columns:
            if ('Where did you' in col) and row.get(col) == 1:
                locs.append(col.split('[')[1].rstrip(']'))
        other = ss(row.get('If other, please explain',''))
        if other: locs.append(other[:40])
        return ', '.join(locs[:3]) if locs else ''

    out_q3['location'] = out_q3.apply(out_location, axis=1)
    unassigned_out = out_q3[out_q3['coordinator'] == 'Unknown']

    def out_summary(df):
        return df.groupby('coordinator').agg(
            sessions=('date','count'),
            outreach_only=('is_outreach', lambda x: (x & ~df.loc[x.index,'is_canvass']).sum()),
            canvass_only=('is_canvass', lambda x: (x & ~df.loc[x.index,'is_outreach']).sum()),
            both=('is_outreach', lambda x: (x & df.loc[x.index,'is_canvass']).sum()),
            people=('people','sum'),
            out_hrs=('out_hrs','sum'),
            can_hrs=('can_hrs','sum'),
        ).reset_index()

    out_summary_q3 = out_summary(out_q3)

    # ── Circles ───────────────────────────────────────────────────
    # Normalize circles column names — Kintone form uses different names
    col_map = {}
    for col in circles_raw.columns:
        cl = col.lower()
        if 'date of circle' in cl or 'date of training' in cl:
            col_map[col] = 'date_col'
        elif 'name of circle' in cl or 'workshops' in cl or 'trainings' in cl:
            col_map[col] = 'type_col'
        elif 'number of' in cl and ('attendee' in cl or 'circle attendee' in cl or 'all attendee' in cl):
            col_map[col] = 'att_col'
        elif 'first and last name' in cl or 'attendee name' in cl:
            col_map[col] = 'name_col'
        elif 'zip' in cl and ('class' in cl or 'circle' in cl or 'school' in cl):
            col_map[col] = 'zip_col'
    circles_raw['_date']  = pd.to_datetime(circles_raw.get(
        next((c for c,v in col_map.items() if v=='date_col'), None), pd.Series(dtype=str)), errors='coerce') if any(v=='date_col' for v in col_map.values()) else pd.NaT
    circles_raw['_type']  = circles_raw.get(next((c for c,v in col_map.items() if v=='type_col'), None), pd.Series([''] * len(circles_raw)))
    circles_raw['_att']   = circles_raw.get(next((c for c,v in col_map.items() if v=='att_col'), None), pd.Series([''] * len(circles_raw)))
    circles_raw['_names'] = circles_raw.get(next((c for c,v in col_map.items() if v=='name_col'), None), pd.Series([''] * len(circles_raw)))
    circles_raw['_zip']   = circles_raw.get(next((c for c,v in col_map.items() if v=='zip_col'), None), pd.Series([''] * len(circles_raw)))
    circles_raw['date'] = circles_raw['_date']
    circles_raw['coordinator'] = circles_raw['Created by'].apply(email_name)
    circ_q3 = circles_raw[(circles_raw['date'] >= Q_START) & (circles_raw['date'] <= Q_END)].copy()
    circ_sessions = circ_q3.groupby(['date','coordinator','_type']).agg(
        attendees=('_att','first'),
        attendee_names=('_names', lambda x: ', '.join(x.dropna().astype(str).str.strip().replace('',pd.NA).dropna()))
    ).reset_index().rename(columns={'_type': 'Workshops / Trainings'}).sort_values(['coordinator','date'])
    circ_summary = circ_sessions.groupby('coordinator').agg(
        total_events=('date','count'),
        total_attendees=('attendees', lambda x: pd.to_numeric(x, errors='coerce').fillna(0).sum())
    ).reset_index()

    # ── Incidents ─────────────────────────────────────────────────
    incidents_raw['date'] = pd.to_datetime(incidents_raw['Date Of Incident Started'], errors='coerce')
    inc_q3 = incidents_raw[(incidents_raw['date'] >= Q_START) & (incidents_raw['date'] <= Q_END)].copy()
    coord_cols_inc = [col for col in incidents_raw.columns if 'Coordinators who were involved' in col]

    def inc_coords(row):
        names = []
        for col in coord_cols_inc:
            if row.get(col) == 1:
                full = col.split('[')[1].rstrip(']')
                names.append(INC_NAME_MAP.get(full, full))
        return ', '.join(names)

    inc_q3['coord_names'] = inc_q3.apply(inc_coords, axis=1)

    # ── Corrections ───────────────────────────────────────────────
    corrections = []

    for _, r in q3_inelig.iterrows():
        corrections.append(['Ineligible: Age Outside 18–40','Participant Folders',r['coordinator'],r['pid'],
            f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))} (Age: {ss(r.get('Current Age',''))})",
            'Exclude from PMRQ. Verify age in Kintone or move to Non-Participant.'])

    for _, r in q3_elig[q3_elig['total_goals'] == 0].iterrows():
        corrections.append(['Missing Goals','Goals Form',r['coordinator'],r['pid'],
            f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))}",
            'No goals found. Add at least one goal in Kintone.'])

    for _, r in q3_elig[~q3_elig['pid'].isin(pre_pids_all)].iterrows():
        corrections.append(['Missing Pre-Assessment','Pre Assessment Form',r['coordinator'],r['pid'],
            f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))}",
            'No pre-assessment on file. Complete in Kintone.'])

    for _, r in q3_elig[~q3_elig['pid'].isin(post_pids_all)].iterrows():
        corrections.append(['Missing Post-Assessment','Post Assessment Form',r['coordinator'],r['pid'],
            f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))}",
            'Complete when participant exits program.'])

    for _, r in q3_elig.iterrows():
        if attest_display(r['pid']) == '⚠ Missing':
            corrections.append(['Missing Self-Attestation','Participant Folders',r['coordinator'],r['pid'],
                f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))}",
                'Self-attestation not completed. Have participant sign in Kintone.'])

    for _, r in q3_elig.iterrows():
        pid = r['pid']
        folder_val = ss(r.get('Did the participant complete the self attestaion form?',''))
        form_found = False
        name_parts = (ss(r.get('First Name','')) + ' ' + ss(r.get('Last Name',''))).strip().lower().split()
        if name_parts:
            for key in attest_name_map:
                if name_parts[0] in key: form_found = True; break
        if not form_found and pid in attest_pid_map: form_found = True
        if folder_val.lower() == 'yes' and not form_found:
            corrections.append(['Self-Attestation Discrepancy','Participant Folders / Attestation Form',
                r['coordinator'], pid,
                f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))}",
                'Folder checkbox = Yes but no form found. Have participant re-submit or fix name.'])

    for _, r in q3_elig[q3_elig['dual_coord']].iterrows():
        corrections.append(['Dual Coordinator Assigned','Participant Folders',r['coordinator'],r['pid'],
            f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))}",
            'Two coordinators assigned. Assign to one only in Kintone.'])

    if len(unassigned_out) > 0:
        corrections.append(['Unassigned Outreach Sessions','Outreach & Canvass Form','Unknown','',
            f"{len(unassigned_out)} sessions with no coordinator identified",
            "Add coordinator email to 'Created by' or 'Trusted Messengers' in Kintone."])

    zero_circ = circ_sessions[pd.to_numeric(circ_sessions['attendees'], errors='coerce').fillna(0) == 0]
    for _, r in zero_circ.iterrows():
        corrections.append(['Circle — 0 Attendees','Circle Tracking',r['coordinator'],'',
            f"{r['date'].strftime('%m/%d/%Y')} | {r['Workshops / Trainings']}",
            'No attendees logged. Confirm cancelled or add records in Kintone.'])

    for _, r in rollover[rollover['total_goals'] == 0].iterrows():
        corrections.append(['Rollover — Missing Goals','Goals Form',r['coordinator'],r['pid'],
            f"{ss(r.get('First Name',''))} {ss(r.get('Last Name',''))} (Prior quarter — In Progress)",
            'No goals on file. Update or close case if inactive.'])

    all_participant_pids = set(atx_dedup['pid'])
    orphan_goal_pids = set(goals['pid']) - all_participant_pids - {'','nan'}
    for opid in sorted(orphan_goal_pids):
        m = goals[goals['pid'] == opid].iloc[0]
        corrections.append(['Goal — No Participant Record','Goals Form',
            email_name(ss(m.get('Assigned Coordinator',''))), opid,
            f"{ss(m.get('Client First Name',''))} {ss(m.get('Client Last Name',''))}",
            'Goal record exists but no matching participant folder. Verify PID in Kintone.'])

    # ════════════════════════════════════════════════════════════════
    # BUILD WORKBOOK
    # ════════════════════════════════════════════════════════════════
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: PMRQ Progress ────────────────────────────────────
    ws = wb.create_sheet("PMRQ Progress")
    title_row(ws, 1, 1, 6, f"ATX Peace {QUARTER} {FY}  |  PMRQ Progress vs. Target")
    subtitle_row(ws, 2, 1, 6, f"{PERIOD}  (quarter-start cases only)  |  Age {AGE_MIN}–{AGE_MAX}  |  Run: {TODAY}")
    for c, h in enumerate(['Metric','Current','Funder Target','Status'], 1):
        H(ws, 3, c, h, sz=11)
    ws.row_dimensions[3].height = 22
    pr = 4

    def pmrq_row(r, metric, current, target, fmt='number'):
        bg = LGR if r % 2 == 0 else WHT
        cell(ws, r, 1, metric, bg=bg, bold=True)
        vc = ws.cell(row=r, column=2, value=current)
        vc.font = Font(name='Arial', size=10); vc.fill = PatternFill('solid', start_color=bg)
        vc.alignment = Alignment(horizontal='center', vertical='center')
        if fmt == 'pct': vc.number_format = '0.0%'
        tc = ws.cell(row=r, column=3, value=target)
        tc.font = Font(name='Arial', size=10); tc.fill = PatternFill('solid', start_color=bg)
        tc.alignment = Alignment(horizontal='center', vertical='center')
        if fmt == 'pct' and isinstance(target, float): tc.number_format = '0.0%'
        hit = False; on_track = False
        try:
            hit = current >= float(target); on_track = current >= float(target) * 0.85
            status = '✅ On Target' if hit else ('🟡 In Progress' if on_track else '🔴 Needs Attention')
        except: status = '— Target TBD'
        sc = ws.cell(row=r, column=4, value=status)
        sc.font = Font(name='Arial', size=10, bold=hit)
        sc.fill = PatternFill('solid', start_color=('D5F5D5' if (isinstance(hit,bool) and hit)
            else ('FFF9C4' if (isinstance(on_track,bool) and on_track) else 'FFD5D5')))
        sc.alignment = Alignment(horizontal='center', vertical='center')

    pmrq_row(pr, f'{QUARTER} Participants Served — LA Quota Target', Q_QUOTA_TOT, '—'); pr+=1
    pmrq_row(pr, f'  Case-Managed (Age {AGE_MIN}–{AGE_MAX})', len(q3_elig), Q_QUOTA_CM); pr+=1
    pmrq_row(pr, '  Non-Participants (community contacts, not PMRQ counted)', len(q3_nonpart), '—'); pr+=1
    pmrq_row(pr, '  Age-Ineligible (excluded from count)', len(q3_inelig), 0); pr+=2
    pmrq_row(pr, '5B Numerator — Participants Showing Improvement', ls_num, '—'); pr+=1
    pmrq_row(pr, f'5B Denominator — Total {QUARTER} Case-Managed', len(q3_elig), '—'); pr+=1
    pmrq_row(pr, '5B Improvement %', ls_pct, TARGET_5B, fmt='pct'); pr+=2
    pmrq_row(pr, 'Participants with Pre-Assessment on File', len(pre_pids_all & q3_pids), '—'); pr+=1
    pmrq_row(pr, 'Participants with Post-Assessment on File', len(post_pids_all & q3_pids), '—'); pr+=1
    pmrq_row(pr, 'Participants with Self-Attestation', len([p for p in q3_pids if attest_display(p) != '⚠ Missing']), '—'); pr+=2
    pmrq_row(pr, f'{QUARTER} Outreach & Canvass Sessions', len(out_q3), Q_QUOTA_OUT); pr+=1
    pmrq_row(pr, f'{QUARTER} Outreach — People Reached', int(out_q3['people'].sum()), '—'); pr+=1
    pmrq_row(pr, f'{QUARTER} Circle / Class Events', len(circ_sessions), '—'); pr+=1
    pmrq_row(pr, f'{QUARTER} Incident Responses', len(inc_q3), '—'); pr+=2
    pmrq_row(pr, 'Rollover Cases — In Progress from Prior Quarters (NOT PMRQ)', len(rollover), '—'); pr+=2
    pmrq_row(pr, f'Data Quality — Missing Goals ({QUARTER} participants)', int((q3_elig['total_goals']==0).sum()), 0); pr+=1
    pmrq_row(pr, 'Data Quality — Missing Pre-Assessments', len(q3_pids - pre_pids_all), 0); pr+=1
    pmrq_row(pr, 'Data Quality — Missing Post-Assessments', len(q3_pids - post_pids_all), 0); pr+=1
    pmrq_row(pr, 'Data Quality — Outreach Sessions w/ No Coordinator', len(unassigned_out), 0); pr+=1

    TBR(ws, 3, pr-1, 1, 4)
    nc = ws.cell(row=pr+1, column=1, value=f'Note: {QUARTER} runs {PERIOD}. Only participants whose cases started in {QUARTER} are counted in PMRQ. Rollover shown in coordinator sheets.')
    nc.font = Font(name='Arial', italic=True, size=9, color='666666')
    ws.merge_cells(f'A{pr+1}:D{pr+1}')

    # ── YTD History section ───────────────────────────────────────
    pr += 4
    ws.merge_cells(f'A{pr}:F{pr}')
    ytd_hdr = ws.cell(row=pr, column=1, value=f'YEAR-TO-DATE HISTORY — {FY}  (Prior quarters locked — as submitted to City)')
    ytd_hdr.font = Font(name='Arial', bold=True, size=11, color=WHT)
    ytd_hdr.fill = PatternFill('solid', start_color=NAVY)
    ytd_hdr.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[pr].height = 20; pr += 1

    for cc, h in enumerate(['Quarter','Dates','Participants\n(City-submitted)','5B %','Outreach\nSessions','People\nReached'], 1):
        H(ws, pr, cc, h, wrap=True, sz=10)
    ws.row_dimensions[pr].height = 32; pr += 1

    q3_out_sessions = len(out_q3)
    q3_out_people   = int(out_q3['people'].sum())
    ytd_cm_thru_q3  = YTD_Q1['atx_peace'] + YTD_Q1['tyj'] + YTD_Q2['total'] + len(q3_elig)
    ytd_out_thru_q3 = (YTD_Q2['outreach_sessions'] or 0) + q3_out_sessions
    q4_cm_needed    = max(0, ANNUAL_GOAL_CM  - ytd_cm_thru_q3)
    q4_out_needed   = max(0, ANNUAL_GOAL_OUT - ytd_out_thru_q3)
    proj_cm  = int(round(ytd_cm_thru_q3 + ytd_cm_thru_q3 / 3))
    proj_out = int(round(ytd_out_thru_q3 + ytd_out_thru_q3 / 2))

    ytd_first_row = pr
    ytd_quarter_data = [
        [YTD_Q1['label'], YTD_Q1['dates'],
         f"{YTD_Q1['total']} ({YTD_Q1['atx_peace']} ATX Peace + {YTD_Q1['tyj']} TYJ)",
         f"{YTD_Q1['5b_pct']:.1%}", '(events-based)', '—'],
        [YTD_Q2['label'], YTD_Q2['dates'],
         f"{YTD_Q2['total']} ({YTD_Q2['atx_peace']} ATX Peace + {YTD_Q2['tyj']} TYJ)",
         f"{YTD_Q2['5b_pct']:.1%}", str(YTD_Q2['outreach_sessions']), str(YTD_Q2['people_reached'])],
        [QUARTER + ' (live)', PERIOD, f"{len(q3_elig)} ATX Peace (+ TYJ TBD)",
         f"{ls_pct:.1%}", str(q3_out_sessions), str(q3_out_people)],
        ['Q4', 'Jul 1 – Sep 30, 2026', '—', '—', '—', '—'],
    ]
    for i, row_data in enumerate(ytd_quarter_data):
        is_live = i == 2; is_locked = i < 2
        bg_row = TEAL if is_live else (GRN if is_locked else LGRY)
        for cc, v in enumerate(row_data, 1):
            cl = ws.cell(row=pr, column=cc, value=v)
            cl.font = Font(name='Arial', size=10, bold=is_live, italic=is_locked)
            cl.fill = PatternFill('solid', start_color=bg_row)
            cl.alignment = Alignment(horizontal='center' if cc > 1 else 'left', vertical='center')
        pr += 1

    # Totals row
    for cc, v in enumerate(['YTD TOTAL (Q1–Q3)','Through June 30, 2026',
            f'{ytd_cm_thru_q3} (Q1–Q3)','—',
            f'{ytd_out_thru_q3} (Q2+Q3; Q1 n/a)','—'], 1):
        cl = ws.cell(row=pr, column=cc, value=v)
        cl.font = Font(name='Arial', size=10, bold=True, color=WHT)
        cl.fill = PatternFill('solid', start_color='2E4057')
        cl.alignment = Alignment(horizontal='center' if cc > 1 else 'left', vertical='center')
    ws.row_dimensions[pr].height = 16; pr += 1

    # Annual goal row
    for cc, v in enumerate([f'Annual Goal ({FY})','Oct 1, 2025 – Sep 30, 2026',
            f'{ANNUAL_GOAL_CM} case-managed', f'{IMPROVEMENT_RATE_TARGET:.0%}',
            f'{ANNUAL_GOAL_OUT} sessions','—'], 1):
        cl = ws.cell(row=pr, column=cc, value=v)
        cl.font = Font(name='Arial', size=10, bold=True)
        cl.fill = PatternFill('solid', start_color=LBLUE)
        cl.alignment = Alignment(horizontal='center' if cc > 1 else 'left', vertical='center')
    ws.row_dimensions[pr].height = 16; pr += 1

    # Q4 needed row
    cm_status  = '✅ Goal met!' if q4_cm_needed  == 0 else f'Need {q4_cm_needed} more in Q4'
    out_status = '✅ Goal met!' if q4_out_needed == 0 else f'Need {q4_out_needed} more in Q4'
    for cc, v in enumerate(['Still Needed in Q4','Jul 1 – Sep 30, 2026',
            cm_status,'—', out_status,'—'], 1):
        hit = '✅' in str(v)
        cl = ws.cell(row=pr, column=cc, value=v)
        cl.font = Font(name='Arial', size=10, bold=True, color='228B22' if hit else '8B0000')
        cl.fill = PatternFill('solid', start_color=GRN if hit else ORG)
        cl.alignment = Alignment(horizontal='center' if cc > 1 else 'left', vertical='center')
    ws.row_dimensions[pr].height = 16; pr += 1

    # Projected year-end row
    proj_cm_txt  = f'~{proj_cm} projected  ({"✅ On pace" if proj_cm >= ANNUAL_GOAL_CM else "⚠ Below pace"})'
    proj_out_txt = f'~{proj_out} projected  ({"✅ On pace" if proj_out >= ANNUAL_GOAL_OUT else "⚠ Below pace"})'
    for cc, v in enumerate(['Projected Year-End (at current pace)','Based on Q1–Q3 avg',
            proj_cm_txt,'—', proj_out_txt,'—'], 1):
        on_pace = '✅' in str(v)
        cl = ws.cell(row=pr, column=cc, value=v)
        cl.font = Font(name='Arial', size=10, italic=True, color='228B22' if on_pace else '8B0000')
        cl.fill = PatternFill('solid', start_color='F0F8F0' if on_pace else 'FFF8E1')
        cl.alignment = Alignment(horizontal='center' if cc > 1 else 'left', vertical='center')
    ws.row_dimensions[pr].height = 16; pr += 1

    TBR(ws, ytd_first_row - 1, pr - 1, 1, 6)
    ln = ws.cell(row=pr, column=1, value='🔒 Q1 and Q2 rows are LOCKED — numbers submitted to City. Do not recalculate from CSVs. Q3 (teal) is live.')
    ln.font = Font(name='Arial', italic=True, size=9, color='8B0000')
    ln.fill = PatternFill('solid', start_color=ORG)
    ws.merge_cells(f'A{pr}:F{pr}')

    for col, w in zip('ABCDEF', [46, 24, 32, 10, 22, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

    # ── Sheet 2: Quota Tracker ────────────────────────────────────
    ws_qt = wb.create_sheet("Quota Tracker")
    title_row(ws_qt, 1, 1, 10, f"ATX Peace {QUARTER} {FY} — Case Management Quota Tracker")
    subtitle_row(ws_qt, 2, 1, 10, f"LA {QUARTER} Target: {Q_LA_TARGET} new case-managed participants")
    qr = 4
    headers = ['Coordinator','M1\nTarget','M1\nActual','M2\nTarget','M2\nActual','M3\nTarget','M3\nActual',
               f'{QUARTER}\nTarget',f'{QUARTER}\nActual','Status']
    for cc, h in enumerate(headers, 1): H(ws_qt, qr, cc, h, wrap=True, sz=10)
    ws_qt.row_dimensions[qr].height = 32; qr += 1

    actuals = {coord: {m: 0 for m in MONTHS} for coord in QUOTA}
    for _, r in q3_elig.iterrows():
        mo = MONTH_MAP.get(r['case_start'].month) if pd.notna(r.get('case_start')) else None
        if mo and r['coordinator'] in actuals: actuals[r['coordinator']][mo] += 1

    tgt_totals = {mo: sum(QUOTA[co][mo] for co in QUOTA) for mo in MONTHS}
    act_totals  = {mo: sum(actuals[co][mo] for co in actuals) for mo in MONTHS}
    total_act_q = sum(act_totals.values())

    for coord, q in QUOTA.items():
        is_lead = coord == 'K. Young'
        act = actuals[coord]
        q_tgt = sum(q.values()); q_act = sum(act.values())
        behind = q_tgt - q_act
        status = ('Lead — counts toward total' if is_lead else
                  '✅ On Track' if q_act >= q_tgt else f'🔴 Needs {behind} more')
        row = [coord, q[MONTHS[0]] or '—', act[MONTHS[0]],
               q[MONTHS[1]] or '—',  act[MONTHS[1]],
               q[MONTHS[2]],         act[MONTHS[2]],
               q_tgt or '—', q_act, status]
        bg = LGR if qr % 2 == 0 else WHT
        for cc, v in enumerate(row, 1):
            cl = ws_qt.cell(row=qr, column=cc, value=v)
            cl.font = Font(name='Arial', size=10, italic=is_lead)
            cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
        for mi, mo in enumerate(MONTHS):
            act_col = 3 + mi * 2
            tgt_val = q[mo]; act_val = act[mo]
            acl = ws_qt.cell(row=qr, column=act_col)
            if tgt_val > 0 and act_val < tgt_val: acl.fill = PatternFill('solid', start_color=RBKG)
            elif tgt_val > 0 and act_val >= tgt_val: acl.fill = PatternFill('solid', start_color=GRN)
        st_cl = ws_qt.cell(row=qr, column=10)
        if '✅' in status: st_cl.fill = PatternFill('solid', start_color=GRN)
        elif '🔴' in status: st_cl.fill = PatternFill('solid', start_color=RBKG)
        else: st_cl.fill = PatternFill('solid', start_color=TEAL)
        qr += 1

    total_q_tgt = sum(sum(QUOTA[co].values()) for co in QUOTA)
    total_row = ['TEAM TOTAL', tgt_totals[MONTHS[0]], act_totals[MONTHS[0]],
                 tgt_totals[MONTHS[1]], act_totals[MONTHS[1]],
                 tgt_totals[MONTHS[2]], act_totals[MONTHS[2]],
                 total_q_tgt, total_act_q,
                 f'{"✅" if total_act_q >= Q_LA_TARGET else "🔴"} {total_act_q} of {Q_LA_TARGET} LA target']
    for cc, v in enumerate(total_row, 1):
        cl = ws_qt.cell(row=qr, column=cc, value=v)
        cl.font = Font(name='Arial', bold=True, size=10)
        cl.fill = PatternFill('solid', start_color=LBLUE)
        cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
    TBR(ws_qt, 4, qr, 1, 10); qr += 2
    for cc, w in zip(range(1, 11), [20,8,8,8,8,8,8,8,8,28]):
        ws_qt.column_dimensions[get_column_letter(cc)].width = w
    ws_qt.freeze_panes = 'A5'

    # ── Sheet 3: Lead Overview ────────────────────────────────────
    ws2 = wb.create_sheet("Lead Overview")
    title_row(ws2, 1, 1, 14, f"ATX Peace {QUARTER} {FY} — Lead Overview by Coordinator")
    subtitle_row(ws2, 2, 1, 14, f"{QUARTER} PMRQ counts only. Rollover shown in coordinator sheets.")
    hdrs = ['Coordinator',f'{QUARTER}\nParticipants','Rollover\n(Review)','Non-\nParticipants',
            'Goals\nMissing ⚠','Pre-Assess\nMissing ⚠','Post-Assess\nMissing ⚠','Attestation\nMissing ⚠',
            'Outreach\nSessions','People\nReached','Out Hrs','Canvass Hrs','Circle\nEvents','Incidents']
    for cc, h in enumerate(hdrs, 1): H(ws2, 3, cc, h, wrap=True)
    ws2.row_dimensions[3].height = 42; dr = 4

    def coord_out_stats(coord):
        g = out_q3[out_q3['coordinator'] == coord]
        return len(g), int(g['people'].sum()), g['out_hrs'].sum(), g['can_hrs'].sum()

    all_display_coords = sorted(set(
        q3_elig['coordinator'].tolist() + rollover['coordinator'].tolist() +
        out_q3[out_q3['coordinator'] != 'Unknown']['coordinator'].tolist() +
        (circ_summary['coordinator'].tolist() if not circ_summary.empty else []) +
        [cc for sub in inc_q3['coord_names'].str.split(', ').dropna() for cc in sub if cc and cc != 'Unknown']
    ))

    for coord in all_display_coords:
        q3c   = q3_elig[q3_elig['coordinator'] == coord]
        rolc  = rollover[rollover['coordinator'] == coord]
        nonpc = q3_nonpart[q3_nonpart['coordinator'] == coord] if 'coordinator' in q3_nonpart.columns else pd.DataFrame()
        mg    = int((q3c['total_goals'] == 0).sum())
        mp    = len(set(q3c['pid']) - pre_pids_all)
        mpo   = len(set(q3c['pid']) - post_pids_all)
        mat   = sum(1 for p in q3c['pid'] if attest_display(p) == '⚠ Missing')
        out_s, out_p, out_o, out_c = coord_out_stats(coord)
        circ_e = len(circ_sessions[circ_sessions['coordinator'] == coord])
        inc_c  = sum(1 for _, r in inc_q3.iterrows() if coord in r['coord_names'])
        bg = LGR if dr % 2 == 0 else WHT
        vals = [coord, len(q3c), len(rolc), len(nonpc), mg, mp, mpo, mat,
                out_s, int(out_p), round(out_o,1), round(out_c,1), circ_e, inc_c]
        for cc, v in enumerate(vals, 1):
            cl = ws2.cell(row=dr, column=cc, value=v)
            cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
            if cc in [5,6,7,8] and isinstance(v,int) and v > 0:
                cl.fill = PatternFill('solid', start_color=ORG)
                cl.font = Font(name='Arial', size=10, bold=True, color='8B0000')
            if cc == 2 and isinstance(v,int) and v > 0: cl.fill = PatternFill('solid', start_color=TEAL)
            if cc == 3 and isinstance(v,int) and v > 0: cl.fill = PatternFill('solid', start_color=ROLLOVER_BG)
        dr += 1

    unrow = ['⚠ UNASSIGNED (no coordinator on record)', 0, 0, 0, 0, 0, 0, 0,
             len(unassigned_out), int(unassigned_out['people'].sum()),
             round(unassigned_out['out_hrs'].sum(),1), round(unassigned_out['can_hrs'].sum(),1), 0, 0]
    for cc, v in enumerate(unrow, 1):
        cl = ws2.cell(row=dr, column=cc, value=v)
        cl.font = Font(name='Arial', size=10, bold=(cc==1), color='8B0000')
        cl.fill = PatternFill('solid', start_color=RBKG)
        cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
    dr += 1
    for cc in range(1, 15):
        val = 'TOTAL' if cc==1 else f'=SUM({get_column_letter(cc)}4:{get_column_letter(cc)}{dr-1})'
        cl = ws2.cell(row=dr, column=cc, value=val)
        cl.font = Font(name='Arial', bold=True, size=10); cl.fill = PatternFill('solid', start_color=LBLUE)
        cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
    TBR(ws2, 3, dr, 1, 14)
    for cc, w in enumerate([22,10,10,10,10,10,10,10,10,10,8,10,8,8], 1):
        ws2.column_dimensions[get_column_letter(cc)].width = w
    ws2.freeze_panes = 'B4'

    # ── Sheet 4: Master Participant List ──────────────────────────
    ws3 = wb.create_sheet("Master Participant List")
    title_row(ws3, 1, 1, 19, f"ATX Peace {QUARTER} {FY} — Master Participant List")
    subtitle_row(ws3, 2, 1, 19, f"{QUARTER} PMRQ participants + Rollover review section")
    mh = ['PID','First Name','Last Name','Age','Age Band','Gender','Race (PMRQ)','Ethnicity','Zip',
          'Coordinator','Source','Case Start','Total Goals','Completed Goals','In Progress',
          'Pre-Assess','Post-Assess','Self-Attestation','Flags']
    for cc, h in enumerate(mh, 1): H(ws3, 3, cc, h, wrap=True)
    ws3.row_dimensions[3].height = 32; mr = 4

    ws3.merge_cells(f'A{mr}:S{mr}')
    sh = ws3.cell(row=mr, column=1, value=f'{QUARTER} PMRQ PARTICIPANTS — Cases started {PERIOD} ({len(q3_elig)} eligible)')
    sh.font = Font(name='Arial', bold=True, size=11, color=WHT)
    sh.fill = PatternFill('solid', start_color=DARK_GRN)
    sh.alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[mr].height = 18; mr += 1

    for _, r in q3_elig.sort_values(['coordinator','Last Name']).iterrows():
        flags = []
        if r['total_goals'] == 0: flags.append('⚠ No goals')
        if r['pid'] not in pre_pids_all: flags.append('⚠ No pre-assess')
        if r['pid'] not in post_pids_all: flags.append('⚠ No post-assess')
        if attest_display(r['pid']) == '⚠ Missing': flags.append('⚠ No attestation')
        if r['dual_coord']: flags.append('⚠ Dual coordinator')
        bg = TEAL if not flags else (LGR if mr % 2 == 0 else WHT)
        row_data = [r['pid'], ss(r.get('First Name','')), ss(r.get('Last Name','')),
            ss(r.get('Current Age','')), map_age_band(r.get('Current Age','')),
            ss(r.get('Gender Identity','')), map_race(r.get('Race','')), map_eth(r.get('Race','')),
            ss(r.get('Zip','')), r['coordinator'], f'{QUARTER} Participant',
            r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
            r['total_goals'], r['completed'], r['in_progress'],
            '✓' if r['pid'] in pre_pids_all else '⚠ Missing',
            '✓' if r['pid'] in post_pids_all else '⚠ Missing',
            attest_display(r['pid']),
            ' | '.join(flags) if flags else '✓ Complete']
        for cc, v in enumerate(row_data, 1):
            cl = ws3.cell(row=mr, column=cc, value=v)
            cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==19))
            if cc==19 and '⚠' in str(v): cl.fill = PatternFill('solid', start_color=ORG)
            if cc in [16,17] and '⚠' in str(v): cl.fill = PatternFill('solid', start_color=LBLUE)
        mr += 1

    if not q3_inelig.empty:
        mr += 1; ws3.merge_cells(f'A{mr}:S{mr}')
        sh2 = ws3.cell(row=mr, column=1, value=f'{QUARTER} AGE-INELIGIBLE — Excluded from PMRQ ({len(q3_inelig)})')
        sh2.font = Font(name='Arial', bold=True, size=11, color=WHT)
        sh2.fill = PatternFill('solid', start_color=DRED)
        sh2.alignment = Alignment(horizontal='left', vertical='center')
        ws3.row_dimensions[mr].height = 18; mr += 1
        for _, r in q3_inelig.iterrows():
            for cc, v in enumerate([r['pid'], ss(r.get('First Name','')), ss(r.get('Last Name','')),
                    ss(r.get('Current Age','')), map_age_band(r.get('Current Age','')),
                    '','','','', r['coordinator'], 'INELIGIBLE',
                    r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
                    0, 0, 0, '', '', '', '⚠ Age outside range. Exclude from PMRQ.'], 1):
                cl = ws3.cell(row=mr, column=cc, value=v)
                cl.font = Font(name='Arial', size=10, bold=True, color=DRED)
                cl.fill = PatternFill('solid', start_color=RBKG)
                cl.alignment = Alignment(horizontal='left', vertical='center')
            mr += 1

    if not q3_nonpart.empty:
        mr += 1; ws3.merge_cells(f'A{mr}:S{mr}')
        sh3 = ws3.cell(row=mr, column=1, value=f'{QUARTER} NON-PARTICIPANTS (Community Contacts — Not PMRQ) ({len(q3_nonpart)})')
        sh3.font = Font(name='Arial', bold=True, size=11, color=WHT)
        sh3.fill = PatternFill('solid', start_color='7B3F8C')
        sh3.alignment = Alignment(horizontal='left', vertical='center')
        ws3.row_dimensions[mr].height = 18; mr += 1
        for _, r in q3_nonpart.iterrows():
            for cc, v in enumerate([r['pid'], ss(r.get('First Name','')), ss(r.get('Last Name','')),
                    ss(r.get('Current Age','')), '', '', '', '', ss(r.get('Zip','')),
                    r['coordinator'], 'Non-Participant',
                    r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
                    'N/A','N/A','N/A','N/A','N/A','N/A','Community contact — not enrolled.'], 1):
                cl = ws3.cell(row=mr, column=cc, value=v)
                cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=LAVENDER)
                cl.alignment = Alignment(horizontal='left', vertical='center')
            mr += 1

    mr += 1; ws3.merge_cells(f'A{mr}:S{mr}')
    sh4 = ws3.cell(row=mr, column=1, value=f'ROLLOVER — In Progress from Prior Quarters ({len(rollover)}) — NOT {QUARTER} PMRQ')
    sh4.font = Font(name='Arial', bold=True, size=11, color=WHT)
    sh4.fill = PatternFill('solid', start_color='2E75B6')
    sh4.alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[mr].height = 18; mr += 1

    for _, r in rollover.sort_values(['coordinator','Last Name']).iterrows():
        flags = []
        if r['total_goals'] == 0: flags.append('⚠ No goals — review or close')
        row_data = [r['pid'], ss(r.get('First Name','')), ss(r.get('Last Name','')),
            ss(r.get('Current Age','')), map_age_band(r.get('Current Age','')),
            ss(r.get('Gender Identity','')), map_race(r.get('Race','')), map_eth(r.get('Race','')),
            ss(r.get('Zip','')), r['coordinator'], 'Rollover (In Progress)',
            r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
            r['total_goals'], r['completed'], r['in_progress'],
            '✓' if r['pid'] in pre_pids_all else '—',
            '✓' if r['pid'] in post_pids_all else '—',
            attest_display(r['pid']),
            ' | '.join(flags) if flags else 'Review — is case still active?']
        for cc, v in enumerate(row_data, 1):
            cl = ws3.cell(row=mr, column=cc, value=v)
            cl.font = Font(name='Arial', size=10, color='444444')
            cl.fill = PatternFill('solid', start_color=ROLLOVER_BG)
            cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==19))
            if cc==19 and '⚠' in str(v): cl.fill = PatternFill('solid', start_color=ORG)
        mr += 1

    TBR(ws3, 3, mr-1, 1, 19)
    for cc, w in enumerate([12,14,16,6,12,14,22,20,10,14,16,12,10,10,10,12,12,18,34], 1):
        ws3.column_dimensions[get_column_letter(cc)].width = w
    ws3.freeze_panes = 'C4'

    # ── Sheet 5: Outreach & Canvass ───────────────────────────────
    ws_oc = wb.create_sheet("Outreach & Canvass")
    title_row(ws_oc, 1, 1, 8, f"ATX Peace {QUARTER} {FY} — Outreach & Canvass Activity")
    subtitle_row(ws_oc, 2, 1, 8, f"{PERIOD}  |  {len(out_q3)} sessions  |  {int(out_q3['people'].sum())} people reached")
    SEC(ws_oc, 3, 1, 8, 'COORDINATOR SUMMARY')
    for cc, h in enumerate(['Coordinator','Sessions','Outreach\nOnly','Canvass\nOnly','Both',
                             'People\nReached','Outreach Hrs','Canvass Hrs'], 1):
        H(ws_oc, 4, cc, h, wrap=True)
    ws_oc.row_dimensions[4].height = 32; sr = 5
    for _, row in out_summary_q3.sort_values('sessions', ascending=False).iterrows():
        bg = RBKG if row['coordinator'] == 'Unknown' else (LGR if sr % 2 == 0 else WHT)
        vals = [row['coordinator'], int(row['sessions']), int(row['outreach_only']),
                int(row['canvass_only']), int(row['both']), int(row['people']),
                round(row['out_hrs'],1), round(row['can_hrs'],1)]
        for cc, v in enumerate(vals, 1):
            cl = ws_oc.cell(row=sr, column=cc, value=v)
            cl.font = Font(name='Arial', size=10,
                           bold=(row['coordinator']=='Unknown'),
                           color='8B0000' if row['coordinator']=='Unknown' else '000000')
            cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
        sr += 1
    for cc in range(1, 9):
        val = 'TOTAL' if cc==1 else f'=SUM({get_column_letter(cc)}5:{get_column_letter(cc)}{sr-1})'
        cl = ws_oc.cell(row=sr, column=cc, value=val)
        cl.font = Font(name='Arial', bold=True, size=10); cl.fill = PatternFill('solid', start_color=LBLUE)
        cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
    TBR(ws_oc, 4, sr, 1, 8); sr += 2
    SEC(ws_oc, sr, 1, 8, f'ACTIVITY LOG — ALL {QUARTER} SESSIONS'); sr += 1
    for cc, h in enumerate(['Date','Coordinator','Type','Outreach Hrs','Canvass Hrs','People Reached','Location','Notes'], 1):
        H(ws_oc, sr, cc, h, bg='2E75B6')
    ws_oc.row_dimensions[sr].height = 18; sr += 1
    for _, r in out_q3.sort_values(['coordinator','date']).iterrows():
        is_unassigned = r['coordinator'] == 'Unknown'
        bg = RBKG if is_unassigned else (LGR if sr % 2 == 0 else WHT)
        notes_text = str(r['notes'])[:300] if pd.notna(r['notes']) and str(r['notes']).strip() else ''
        for cc, v in enumerate([
                r['date'].strftime('%m/%d/%Y') if pd.notna(r['date']) else '',
                r['coordinator'], r['out_type'],
                r['out_hrs'] if r['out_hrs'] > 0 else '',
                r['can_hrs'] if r['can_hrs'] > 0 else '',
                int(r['people']) if r['people'] > 0 else '⚠ 0',
                r['location'][:50], notes_text], 1):
            cl = ws_oc.cell(row=sr, column=cc, value=v)
            cl.font = Font(name='Arial', size=9, color='8B0000' if is_unassigned else '000000')
            cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==8))
        ws_oc.row_dimensions[sr].height = max(14, min(60, len(notes_text)//80*14+14))
        sr += 1
    TBR(ws_oc, 4, sr-1, 1, 8)
    for cc, w in zip(range(1,9), [12,14,10,10,10,12,30,60]):
        ws_oc.column_dimensions[get_column_letter(cc)].width = w
    ws_oc.freeze_panes = 'A8'

    # ── Sheet 6: Circles & Classes ────────────────────────────────
    ws_cc = wb.create_sheet("Circles & Classes")
    title_row(ws_cc, 1, 1, 7, f"ATX Peace {QUARTER} {FY} — Circles & Classes")
    subtitle_row(ws_cc, 2, 1, 7, f"{QUARTER}: {len(circ_sessions)} sessions")
    SEC(ws_cc, 3, 1, 7, 'COORDINATOR SUMMARY')
    for cc, h in enumerate(['Coordinator','Total Events','Total Attendees','Life Skills',
                             'Community Building','Circle of Support','Conflict Circle'], 1):
        H(ws_cc, 4, cc, h, wrap=True)
    ws_cc.row_dimensions[4].height = 32; sr_cc = 5
    for coord in sorted(circ_sessions['coordinator'].unique()):
        cg = circ_sessions[circ_sessions['coordinator'] == coord]
        att = pd.to_numeric(cg['attendees'], errors='coerce').fillna(0).sum()
        def count_type(t): return len(cg[cg['Workshops / Trainings'].str.contains(t, case=False, na=False)])
        bg = LGR if sr_cc % 2 == 0 else WHT
        for cc, v in enumerate([coord, len(cg), int(att),
                count_type('Life Skills'), count_type('Community'),
                count_type('Support'), count_type('Conflict')], 1):
            cl = ws_cc.cell(row=sr_cc, column=cc, value=v)
            cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
        sr_cc += 1
    for cc in range(1, 8):
        val = 'TOTAL' if cc==1 else f'=SUM({get_column_letter(cc)}5:{get_column_letter(cc)}{sr_cc-1})'
        cl = ws_cc.cell(row=sr_cc, column=cc, value=val)
        cl.font = Font(name='Arial', bold=True, size=10); cl.fill = PatternFill('solid', start_color=LBLUE)
        cl.alignment = Alignment(horizontal='left' if cc==1 else 'center', vertical='center')
    TBR(ws_cc, 4, sr_cc, 1, 7); sr_cc += 2
    SEC(ws_cc, sr_cc, 1, 7, f'ALL {QUARTER} EVENTS'); sr_cc += 1
    for cc, h in enumerate(['Date','Coordinator','Type','Location (Zip)','Attendees','Names on Record','Notes'], 1):
        H(ws_cc, sr_cc, cc, h, bg='2E75B6')
    ws_cc.row_dimensions[sr_cc].height = 18; sr_cc += 1
    for _, r in circ_sessions.sort_values(['coordinator','date']).iterrows():
        att_val = pd.to_numeric(r['attendees'], errors='coerce')
        zero_att = pd.isna(att_val) or att_val == 0
        bg = YEL if zero_att else (LGR if sr_cc % 2 == 0 else WHT)
        for cc, v in enumerate([
                r['date'].strftime('%m/%d/%Y') if pd.notna(r['date']) else '',
                r['coordinator'], r['Workshops / Trainings'],
                ss(r.get('_zip', '')),
                int(att_val) if pd.notna(att_val) else '⚠ 0',
                r['attendee_names'][:80] if r['attendee_names'] else '',
                '⚠ 0 attendees' if zero_att else ''], 1):
            cl = ws_cc.cell(row=sr_cc, column=cc, value=v)
            cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc in [6,7]))
        sr_cc += 1
    TBR(ws_cc, 4, sr_cc-1, 1, 7)
    for cc, w in zip(range(1,8), [12,14,22,14,10,40,20]):
        ws_cc.column_dimensions[get_column_letter(cc)].width = w
    ws_cc.freeze_panes = 'A8'

    # ── Sheet 7: Incidents ────────────────────────────────────────
    ws_inc = wb.create_sheet(f"Incidents — {QUARTER}")
    title_row(ws_inc, 1, 1, 7, f"ATX Peace {QUARTER} {FY} — Incident Responses", bg='722F37')
    subtitle_row(ws_inc, 2, 1, 7, f"{QUARTER}: {len(inc_q3)} incidents")
    for cc, h in enumerate(['Date','Type','Address','Zip','Coordinators Involved','Technique Used','Conflict Description'], 1):
        H(ws_inc, 3, cc, h, bg='722F37', wrap=True)
    ws_inc.row_dimensions[3].height = 28; ir = 4
    for _, r in inc_q3.sort_values('date').iterrows():
        is_shooting = 'Shooting' in str(r.get('Type of Incident',''))
        bg = RBKG if is_shooting else INC_BG
        desc = ss(r.get('Conflict Description',''))[:400]
        for cc, v in enumerate([
                r['date'].strftime('%m/%d/%Y') if pd.notna(r['date']) else '',
                ss(r.get('Type of Incident','')), ss(r.get('Address Line 1','')),
                ss(r.get('Zipcode','')), r['coord_names'],
                ss(r.get('Techniques Used','')), desc], 1):
            cl = ws_inc.cell(row=ir, column=cc, value=v)
            cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==7))
        ws_inc.row_dimensions[ir].height = max(14, min(80, len(desc)//70*14+14)); ir += 1
    TBR(ws_inc, 3, ir-1, 1, 7)
    for cc, w in zip(range(1,8), [12,24,24,8,28,22,60]):
        ws_inc.column_dimensions[get_column_letter(cc)].width = w
    ws_inc.freeze_panes = 'A4'

    # ── Sheets 8+: Per Coordinator ────────────────────────────────
    for coord in all_display_coords:
        short = coord.replace('. ','_').replace(' ','_')[:20]
        wsc = wb.create_sheet(f"Coord_{short}")
        title_row(wsc, 1, 1, 14, f"ATX Peace {QUARTER} — {coord}  |  Updated {TODAY}")
        cr = 2

        SEC(wsc, cr, 1, 14,
            f"PRIOR & CURRENT WEEK  "
            f"({PREV_WEEK_START.strftime('%b %-d')}–{PREV_WEEK_END.strftime('%-d')}  /  "
            f"{WEEK_START.strftime('%b %-d')}–{WEEK_END.strftime('%-d, %Y')})",
            bg='1A5276'); cr += 1

        def week_label(start, end, label):
            return f"{label}  ({start.strftime('%b %-d')}–{end.strftime('%-d')})"

        def write_week_rows(wsc, cr, label, fu_rows, out_rows, circ_rows, week_bg):
            """Write one week's block of follow-ups / outreach / circles."""
            # Sub-header
            for col in range(1, 15):
                cl = wsc.cell(row=cr, column=col, value=label if col == 1 else '')
                cl.font = Font(name='Arial', bold=True, size=9, color=WHT)
                cl.fill = PatternFill('solid', start_color=week_bg)
                cl.alignment = Alignment(horizontal='left', vertical='center')
            wsc.merge_cells(f'A{cr}:N{cr}'); cr += 1

            # Follow-ups
            wsc.cell(row=cr, column=1, value='Follow-ups').font = Font(name='Arial', bold=True, size=10)
            wsc.cell(row=cr, column=1).fill = PatternFill('solid', start_color=LBLUE)
            wsc.cell(row=cr, column=1).alignment = Alignment(horizontal='left', vertical='center'); cr += 1
            if fu_rows.empty:
                cl = wsc.cell(row=cr, column=1, value='No follow-ups.')
                cl.font = Font(name='Arial', size=10, italic=True, color='666666')
                cl.fill = PatternFill('solid', start_color=LGR)
                wsc.merge_cells(f'A{cr}:N{cr}'); cr += 1
            else:
                for _, r in fu_rows.sort_values('date').iterrows():
                    conn = r.get('Did you make connection with the participant?','')
                    conn_icon = '✅' if conn=='Connected' else ('❌' if pd.notna(conn) and conn else '—')
                    name  = f"{ss(r.get('Participant First Name',''))} {ss(r.get('Participant Last Name',''))}".strip()
                    date_s = r['date'].strftime('%m/%d') if pd.notna(r['date']) else ''
                    support  = ss(r.get('What support did you give during this follow up?',''))
                    duration = ss(r.get('How long was follow up?',''))
                    summary  = f"{date_s}  |  {name}  |  {conn_icon}  |  {support}  |  {duration}"
                    cl = wsc.cell(row=cr, column=1, value=summary)
                    cl.font = Font(name='Arial', size=10)
                    cl.fill = PatternFill('solid', start_color=GRN if conn=='Connected' else (RBKG if (pd.notna(conn) and conn) else LGR))
                    cl.alignment = Alignment(horizontal='left', vertical='center')
                    wsc.merge_cells(f'A{cr}:N{cr}'); wsc.row_dimensions[cr].height = 16; cr += 1

            # Outreach & Canvass
            wsc.cell(row=cr, column=1, value='Outreach & Canvass').font = Font(name='Arial', bold=True, size=10)
            wsc.cell(row=cr, column=1).fill = PatternFill('solid', start_color=LBLUE)
            wsc.cell(row=cr, column=1).alignment = Alignment(horizontal='left', vertical='center'); cr += 1
            if out_rows.empty:
                cl = wsc.cell(row=cr, column=1, value='No outreach or canvass.')
                cl.font = Font(name='Arial', size=10, italic=True, color='666666')
                cl.fill = PatternFill('solid', start_color=LGR)
                wsc.merge_cells(f'A{cr}:N{cr}'); cr += 1
            else:
                for _, r in out_rows.sort_values('date').iterrows():
                    date_s = r['date'].strftime('%m/%d') if pd.notna(r['date']) else ''
                    people = int(r['people']) if r['people'] > 0 else '0 logged'
                    summary = f"{date_s}  |  {r['out_type']}  |  {r['out_hrs']:.0f}h out / {r['can_hrs']:.0f}h can  |  {people} people"
                    cl = wsc.cell(row=cr, column=1, value=summary)
                    cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=TEAL)
                    cl.alignment = Alignment(horizontal='left', vertical='center')
                    wsc.merge_cells(f'A{cr}:N{cr}'); wsc.row_dimensions[cr].height = 16; cr += 1

            # Circles & Classes
            wsc.cell(row=cr, column=1, value='Circles & Classes').font = Font(name='Arial', bold=True, size=10)
            wsc.cell(row=cr, column=1).fill = PatternFill('solid', start_color=LBLUE)
            wsc.cell(row=cr, column=1).alignment = Alignment(horizontal='left', vertical='center'); cr += 1
            if circ_rows.empty:
                cl = wsc.cell(row=cr, column=1, value='No circles or classes.')
                cl.font = Font(name='Arial', size=10, italic=True, color='666666')
                cl.fill = PatternFill('solid', start_color=LGR)
                wsc.merge_cells(f'A{cr}:N{cr}'); cr += 1
            else:
                for _, r in circ_rows.sort_values('date').iterrows():
                    date_s = r['date'].strftime('%m/%d') if pd.notna(r['date']) else ''
                    att = pd.to_numeric(r['attendees'], errors='coerce')
                    names = r['attendee_names'][:80] if r['attendee_names'] else ''
                    summary = f"{date_s} | {r['Workshops / Trainings']} | {int(att) if pd.notna(att) else 0} attendees" + (f" — {names}" if names else '')
                    cl = wsc.cell(row=cr, column=1, value=summary)
                    cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=PURPLE)
                    cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    wsc.merge_cells(f'A{cr}:N{cr}'); wsc.row_dimensions[cr].height = 14; cr += 1
            return cr

        # Prior week
        cr = write_week_rows(
            wsc, cr,
            week_label(PREV_WEEK_START, PREV_WEEK_END, "PRIOR WEEK"),
            fu_q3[(fu_q3['date'] >= PREV_WEEK_START) & (fu_q3['date'] <= PREV_WEEK_END) &
                  (fu_q3['Assigned Coordinator'].apply(email_name) == coord)],
            out_q3[(out_q3['date'] >= PREV_WEEK_START) & (out_q3['date'] <= PREV_WEEK_END) &
                   (out_q3['coordinator'] == coord)],
            circ_sessions[(circ_sessions['date'] >= PREV_WEEK_START) & (circ_sessions['date'] <= PREV_WEEK_END) &
                          (circ_sessions['coordinator'] == coord)],
            week_bg='2471A3',
        )

        # Current week
        cr = write_week_rows(
            wsc, cr,
            week_label(WEEK_START, WEEK_END, "CURRENT WEEK"),
            fu_q3[(fu_q3['date'] >= WEEK_START) & (fu_q3['date'] <= WEEK_END) &
                  (fu_q3['Assigned Coordinator'].apply(email_name) == coord)],
            out_q3[(out_q3['date'] >= WEEK_START) & (out_q3['date'] <= WEEK_END) &
                   (out_q3['coordinator'] == coord)],
            circ_sessions[(circ_sessions['date'] >= WEEK_START) & (circ_sessions['date'] <= WEEK_END) &
                          (circ_sessions['coordinator'] == coord)],
            week_bg='1A5276',
        )

        cr += 1

        # Q3 participants table
        q3c = q3_elig[q3_elig['coordinator'] == coord].sort_values('Last Name')
        last_fu_date = fu_q3.groupby('pid')['date'].max()  # most recent follow-up per PID
        SEC(wsc, cr, 1, 14, f"{QUARTER} CASE-MANAGED PARTICIPANTS ({len(q3c)}) — {QUARTER} PMRQ", bg=DARK_GRN); cr += 1
        for cc, h in enumerate(['PID','First Name','Last Name','Age','Gender','Race','Zip','Case Start',
                                  'Goals','Completed\nGoals','In Prog','Pre','Post','Flags',
                                  'Follow-ups\n(connected/total)','Last\nFollow-up'], 1):
            H(wsc, cr, cc, h, bg=LBLUE, fg='000000', wrap=True, sz=9)
        wsc.row_dimensions[cr].height = 28; cr += 1
        if q3c.empty:
            wsc.cell(row=cr, column=1, value=f'No {QUARTER} participants.').font = Font(name='Arial', italic=True, size=10); cr += 1
        else:
            for _, r in q3c.iterrows():
                pid = r['pid']
                flags = []
                if r['total_goals'] == 0: flags.append('⚠ No goals')
                if pid not in pre_pids_all: flags.append('⚠ No pre-assess')
                if pid not in post_pids_all: flags.append('⚠ No post-assess')
                if attest_display(pid) == '⚠ Missing': flags.append('⚠ No attestation')
                p_fu = fu_q3[fu_q3['pid'] == pid]
                fu_total = len(p_fu)
                fu_conn  = (p_fu['Did you make connection with the participant?'] == 'Connected').sum()
                fu_val   = f"{fu_conn} / {fu_total}"
                last_fu  = last_fu_date.get(pid)
                last_fu_s = last_fu.strftime('%m/%d/%Y') if pd.notna(last_fu) else '—'
                bg = TEAL if not flags else (LGR if cr % 2 == 0 else WHT)
                for cc, v in enumerate([pid, ss(r.get('First Name','')), ss(r.get('Last Name','')),
                        ss(r.get('Current Age','')), ss(r.get('Gender Identity','')),
                        map_race(r.get('Race','')), ss(r.get('Zip','')),
                        r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
                        r['total_goals'], r['completed'], r['in_progress'],
                        '✓' if pid in pre_pids_all else '⚠',
                        '✓' if pid in post_pids_all else '⚠',
                        ' | '.join(flags) if flags else '✓',
                        fu_val, last_fu_s], 1):
                    cl = wsc.cell(row=cr, column=cc, value=v)
                    cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
                    cl.alignment = Alignment(horizontal='center' if cc in [9,10,11,12,13,15,16] else 'left',
                                             vertical='center', wrap_text=(cc==14))
                    if cc in [12,13] and v=='⚠': cl.fill = PatternFill('solid', start_color=LBLUE)
                    if cc==14 and '⚠' in str(v): cl.fill = PatternFill('solid', start_color=ORG)
                    if cc==15 and fu_total==0: cl.fill = PatternFill('solid', start_color=RBKG)
                    if cc==16 and last_fu_s=='—': cl.fill = PatternFill('solid', start_color=RBKG)
                cr += 1

        # Ineligible
        inelig_c = q3_all_inelig[q3_all_inelig['coordinator'] == coord]
        for _, r in inelig_c.iterrows():
            pid = r['pid']
            p_fu_in = fu_q3[fu_q3['pid'] == pid]
            fu_in_conn = (p_fu_in['Did you make connection with the participant?']=='Connected').sum()
            fu_in_val  = f"{fu_in_conn} / {len(p_fu_in)}"
            reason = f"INELIGIBLE — Age {ss(r.get('Current Age','missing'))} (requires {AGE_MIN}–{AGE_MAX})"
            for cc, v in enumerate([pid, ss(r.get('First Name','')), ss(r.get('Last Name','')),
                    ss(r.get('Current Age','')), ss(r.get('Gender Identity','')),
                    map_race(r.get('Race','')), ss(r.get('Zip','')),
                    r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
                    '—','—','—','—','—', reason, fu_in_val], 1):
                cl = wsc.cell(row=cr, column=cc, value=v)
                cl.font = Font(name='Arial', size=10, bold=True, color=DRED)
                cl.fill = PatternFill('solid', start_color=RBKG)
                cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==14))
            cr += 1

        # Rollover
        rolc = rollover[rollover['coordinator'] == coord].sort_values('Last Name')
        cr += 1
        SEC(wsc, cr, 1, 14, f"ROLLOVER — IN PROGRESS FROM PRIOR QUARTERS ({len(rolc)}) — Review for closure", bg='2E75B6'); cr += 1
        for cc, h in enumerate(['PID','First Name','Last Name','Age','Case Start','Goals','Completed\nGoals',
                                  'Goal Categories','Status','Pre','Post','','Flags',
                                  'Follow-ups\n(connected/total)','Last\nFollow-up'], 1):
            H(wsc, cr, cc, h, bg=LBLUE, fg='000000', wrap=True, sz=9)
        wsc.row_dimensions[cr].height = 24; cr += 1
        if rolc.empty:
            wsc.cell(row=cr, column=1, value='No rollover cases.').font = Font(name='Arial', italic=True, size=10); cr += 1
        else:
            for _, r in rolc.iterrows():
                pid = r['pid']
                flags = []
                if r['total_goals'] == 0: flags.append('⚠ No goals — close or update')
                if r['completed'] == 0 and r['in_progress'] == 0: flags.append('No activity')
                p_fu_rol   = fu_q3[fu_q3['pid'] == pid]
                fu_rol_tot = len(p_fu_rol)
                fu_rol_con = (p_fu_rol['Did you make connection with the participant?'] == 'Connected').sum()
                fu_rol_val = f"{fu_rol_con} / {fu_rol_tot}"
                last_fu_rol   = last_fu_date.get(pid)
                last_fu_rol_s = last_fu_rol.strftime('%m/%d/%Y') if pd.notna(last_fu_rol) else '—'
                for cc, v in enumerate([pid, ss(r.get('First Name','')), ss(r.get('Last Name','')),
                        ss(r.get('Current Age','')),
                        r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
                        r['total_goals'], r['completed'], r.get('categories','')[:40],
                        ss(r.get('Case Manage Progress','')),
                        '✓' if pid in pre_pids_all else '—',
                        '✓' if pid in post_pids_all else '—',
                        '', ' | '.join(flags) if flags else 'Active — verify still open',
                        fu_rol_val, last_fu_rol_s], 1):
                    cl = wsc.cell(row=cr, column=cc, value=v)
                    cl.font = Font(name='Arial', size=10, color='444444')
                    cl.fill = PatternFill('solid', start_color=ORG if '⚠' in str(v) else ROLLOVER_BG)
                    cl.alignment = Alignment(horizontal='center' if cc in [6,7,10,11,14,15] else 'left',
                                             vertical='center', wrap_text=(cc==13))
                    if cc==14 and fu_rol_tot==0:
                        cl.fill = PatternFill('solid', start_color=RBKG)
                        cl.font = Font(name='Arial', size=10, bold=True, color=DRED)
                    if cc==15 and last_fu_rol_s=='—':
                        cl.fill = PatternFill('solid', start_color=RBKG)
                cr += 1

        # Outreach log
        out_coord = out_q3[out_q3['coordinator'] == coord].sort_values('date')
        cr += 1
        SEC(wsc, cr, 1, 14, f"OUTREACH & CANVASS — {QUARTER} ({len(out_coord)} sessions | {int(out_coord['people'].sum())} people reached)"); cr += 1
        for cc, h in enumerate(['Total Sessions','Outreach Only','Canvass Only','Both','People Reached','Outreach Hrs','Canvass Hrs','','','','','','',''], 1):
            H(wsc, cr, cc, h, bg=LBLUE, fg='000000', sz=9, wrap=True)
        cr += 1
        stats_row = [len(out_coord),
                     int((out_coord['is_outreach'] & ~out_coord['is_canvass']).sum()),
                     int((out_coord['is_canvass'] & ~out_coord['is_outreach']).sum()),
                     int((out_coord['is_outreach'] & out_coord['is_canvass']).sum()),
                     int(out_coord['people'].sum()), round(out_coord['out_hrs'].sum(),1),
                     round(out_coord['can_hrs'].sum(),1)] + ['']*7
        for cc, v in enumerate(stats_row, 1):
            cell(wsc, cr, cc, v, bg=LGR, align='center' if cc > 1 else 'left')
        cr += 2
        if out_coord.empty:
            wsc.cell(row=cr, column=1, value=f'No outreach/canvass activity in {QUARTER}.').font = Font(name='Arial', italic=True, size=10); cr += 1
        else:
            for cc, h in enumerate(['Date','Type','Outreach Hrs','Canvass Hrs','People Reached','Location','Notes'], 1):
                H(wsc, cr, cc, h, bg='2E75B6', sz=9)
            wsc.row_dimensions[cr].height = 16; cr += 1
            for _, r in out_coord.iterrows():
                notes_text = str(r['notes'])[:250] if str(r['notes']).strip() else ''
                bg = LGR if cr % 2 == 0 else WHT
                for cc, v in enumerate([
                        r['date'].strftime('%m/%d/%Y') if pd.notna(r['date']) else '',
                        r['out_type'],
                        r['out_hrs'] if r['out_hrs'] > 0 else '',
                        r['can_hrs'] if r['can_hrs'] > 0 else '',
                        int(r['people']) if r['people'] > 0 else '⚠ 0',
                        r['location'][:40], notes_text], 1):
                    cl = wsc.cell(row=cr, column=cc, value=v)
                    cl.font = Font(name='Arial', size=9); cl.fill = PatternFill('solid', start_color=bg)
                    cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==7))
                wsc.row_dimensions[cr].height = max(14, min(72, len(notes_text)//70*14+14)); cr += 1

        # Circles log
        circ_coord = circ_sessions[circ_sessions['coordinator'] == coord].sort_values('date')
        cr += 1
        SEC(wsc, cr, 1, 14, f"CIRCLES & CLASSES — {QUARTER} ({len(circ_coord)} events)"); cr += 1
        if circ_coord.empty:
            wsc.cell(row=cr, column=1, value=f'No circles/classes in {QUARTER}.').font = Font(name='Arial', italic=True, size=10); cr += 1
        else:
            for cc, h in enumerate(['Date','Type','Zip','Attendees','Names'], 1):
                H(wsc, cr, cc, h, bg=LBLUE, fg='000000', sz=9)
            cr += 1
            for _, r in circ_coord.iterrows():
                att = pd.to_numeric(r['attendees'], errors='coerce')
                zero = pd.isna(att) or att == 0
                bg = YEL if zero else (LGR if cr % 2 == 0 else WHT)
                for cc, v in enumerate([r['date'].strftime('%m/%d/%Y') if pd.notna(r['date']) else '',
                        r['Workshops / Trainings'], '', int(att) if pd.notna(att) else '⚠ 0',
                        r['attendee_names'][:60]], 1):
                    cl = wsc.cell(row=cr, column=cc, value=v)
                    cl.font = Font(name='Arial', size=9); cl.fill = PatternFill('solid', start_color=bg)
                    cl.alignment = Alignment(horizontal='left', vertical='center')
                cr += 1

        # Incident log
        inc_coord_rows = inc_q3[[coord in r['coord_names'] for _, r in inc_q3.iterrows()]] if coord else pd.DataFrame()
        if not inc_coord_rows.empty:
            cr += 1
            SEC(wsc, cr, 1, 14, f"INCIDENT RESPONSES — {QUARTER} ({len(inc_coord_rows)})", bg='722F37'); cr += 1
            for cc, h in enumerate(['Date','Type','Address','Technique','Description'], 1):
                H(wsc, cr, cc, h, bg='722F37', sz=9)
            cr += 1
            for _, r in inc_coord_rows.iterrows():
                for cc, v in enumerate([r['date'].strftime('%m/%d/%Y') if pd.notna(r['date']) else '',
                        ss(r.get('Type of Incident','')), ss(r.get('Address Line 1','')),
                        ss(r.get('Techniques Used','')), ss(r.get('Conflict Description',''))[:200]], 1):
                    cl = wsc.cell(row=cr, column=cc, value=v)
                    cl.font = Font(name='Arial', size=9)
                    cl.fill = PatternFill('solid', start_color=INC_BG if 'dispute' in str(r.get('Type of Incident','')).lower() else RBKG)
                    cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==5))
                cr += 1

        # Corrections
        coord_corrs = [x for x in corrections if x[2] == coord]
        cr += 1
        SEC(wsc, cr, 1, 14, f"CORRECTIONS NEEDED ({len(coord_corrs)})", bg=DRED); cr += 1
        for cc, h in enumerate(['Issue Type','PID','Name','Detail','Action Required'], 1):
            H(wsc, cr, cc, h, bg=DRED, sz=9)
        cr += 1
        if not coord_corrs:
            wsc.cell(row=cr, column=1, value=f'✓ No corrections needed for {QUARTER} participants.').font = Font(name='Arial', size=10, color='228B22'); cr += 1
        else:
            for corr in coord_corrs:
                bg = corr_color(corr[0])
                for cc, v in enumerate(corr[:5], 1):
                    cl = wsc.cell(row=cr, column=cc, value=v)
                    cl.font = Font(name='Arial', size=9); cl.fill = PatternFill('solid', start_color=bg)
                    cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==5))
                cr += 1

        for cc, w in zip(range(1,15), [12,14,16,6,14,14,10,13,10,8,8,8,12,36]):
            wsc.column_dimensions[get_column_letter(cc)].width = w
        wsc.freeze_panes = 'C3'

    # ── YTD History sheet ─────────────────────────────────────────
    ws_ytd = wb.create_sheet("YTD History")
    title_row(ws_ytd, 1, 1, 7, f"ATX Peace {FY} — Year-to-Date History")
    subtitle_row(ws_ytd, 2, 1, 7, f"Prior quarters LOCKED (as submitted to City) | {QUARTER} is live from current CSV exports")
    ytd_h = ['Quarter','Dates','Total\n(City-submitted)','ATX Peace\nOnly','TYJ','5B %',
             'Outreach\nSessions','People\nReached','Events','Event\nAttendees','Incidents']
    for cc, h in enumerate(ytd_h, 1): H(ws_ytd, 3, cc, h, wrap=True)
    ws_ytd.row_dimensions[3].height = 32

    def ytd_row_fn(ws, rr, data, bg):
        for cc, v in enumerate(data, 1):
            cl = ws.cell(row=rr, column=cc, value=v)
            cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='center' if cc > 1 else 'left', vertical='center')

    yr = 4
    ytd_row_fn(ws_ytd, yr, ['Q1 🔒', YTD_Q1['dates'], YTD_Q1['total'], YTD_Q1['atx_peace'], YTD_Q1['tyj'],
        f"{YTD_Q1['5b_pct']:.1%}", '(events-based)', '—',
        YTD_Q1['events'], YTD_Q1['event_attendees'], YTD_Q1['incidents']], GRN); yr += 1
    ytd_row_fn(ws_ytd, yr, ['Q2 🔒', YTD_Q2['dates'], YTD_Q2['total'], YTD_Q2['atx_peace'], YTD_Q2['tyj'],
        f"{YTD_Q2['5b_pct']:.1%}", YTD_Q2['outreach_sessions'], YTD_Q2['people_reached'],
        YTD_Q2['events'], YTD_Q2['event_attendees'], YTD_Q2['incidents']], GRN); yr += 1
    ytd_row_fn(ws_ytd, yr, [f'{QUARTER} (live)', PERIOD, f'{len(q3_elig)} (ATX Peace)', len(q3_elig), '(TBD)',
        f'{ls_pct:.1%}', len(out_q3), int(out_q3["people"].sum()),
        len(circ_sessions),
        int(circ_sessions["attendees"].apply(lambda x: pd.to_numeric(x, errors="coerce")).fillna(0).sum()),
        len(inc_q3)], TEAL); yr += 1
    ytd_row_fn(ws_ytd, yr, ['Q4', 'Jul 1 – Sep 30, 2026', '—', '—', '—', '—', '—', '—', '—', '—', '—'], LGRY); yr += 1
    TBR(ws_ytd, 3, yr-1, 1, 11); yr += 2

    ws_ytd.merge_cells(f'A{yr}:K{yr}')
    pid_hdr = ws_ytd.cell(row=yr, column=1, value='PARTICIPANT PID LISTS — For cross-quarter duplicate detection')
    pid_hdr.font = Font(name='Arial', bold=True, size=11, color=WHT)
    pid_hdr.fill = PatternFill('solid', start_color=NAVY)
    pid_hdr.alignment = Alignment(horizontal='left', vertical='center')
    ws_ytd.row_dimensions[yr].height = 20; yr += 1

    for qdata, qlabel, pid_key, bg_c in [
        (YTD_Q1, 'Q1 ATX Peace PIDs (locked)', 'pids_atx', GRN),
        (YTD_Q1, 'Q1 TYJ PIDs (locked)', 'pids_tyj', LGRY),
        (YTD_Q2, 'Q2 ATX Peace PIDs (locked)', 'pids_atx', GRN),
        (YTD_Q2, 'Q2 TYJ PIDs (locked)', 'pids_tyj', LGRY),
        (None,   f'{QUARTER} ATX Peace PIDs (live)', None, TEAL),
    ]:
        pid_set = sorted(q3_pids if qdata is None else qdata[pid_key])
        ws_ytd.merge_cells(f'A{yr}:K{yr}')
        lbl_cl = ws_ytd.cell(row=yr, column=1, value=f'{qlabel}:  {", ".join(pid_set)}')
        lbl_cl.font = Font(name='Arial', size=9)
        lbl_cl.fill = PatternFill('solid', start_color=bg_c)
        lbl_cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_ytd.row_dimensions[yr].height = max(14, min(80, len(", ".join(pid_set))//140*14+14)); yr += 1

    for cc, w in zip(range(1,12), [12,22,16,14,8,10,14,14,10,14,10]):
        ws_ytd.column_dimensions[get_column_letter(cc)].width = w
    ws_ytd.freeze_panes = 'B4'

    # ── Duplicate PIDs sheet ──────────────────────────────────────
    ws_dup = wb.create_sheet("Duplicate PIDs")
    title_row(ws_dup, 1, 1, 7, f"ATX Peace {QUARTER} {FY} — Cross-Quarter Duplicate PIDs", bg='C55A11')
    subtitle_row(ws_dup, 2, 1, 7, f"PIDs in {QUARTER} that also appeared in a prior quarter — resolve in Kintone before submission")
    for cc, h in enumerate(['PID','First Name','Last Name','Coordinator','Case Start','Prior Quarter(s)','Action'], 1):
        H(ws_dup, 3, cc, h, bg='C55A11', wrap=True)
    ws_dup.row_dimensions[3].height = 24; dr_dup = 4
    if duplicate_q3.empty:
        cl = ws_dup.cell(row=dr_dup, column=1, value='✅ No cross-quarter duplicate PIDs found.')
        cl.font = Font(name='Arial', size=10, color='228B22')
        cl.fill = PatternFill('solid', start_color=GRN)
        ws_dup.merge_cells(f'A{dr_dup}:G{dr_dup}')
    else:
        for _, r in duplicate_q3.iterrows():
            pid = r['pid']
            prior_q_labels = []
            if pid in (YTD_Q1['pids_atx'] | YTD_Q1['pids_tyj']): prior_q_labels.append('Q1')
            if pid in (YTD_Q2['pids_atx'] | YTD_Q2['pids_tyj']): prior_q_labels.append('Q2')
            for cc, v in enumerate([pid, ss(r.get('First Name','')), ss(r.get('Last Name','')),
                    r['coordinator'],
                    r['case_start'].strftime('%m/%d/%Y') if pd.notna(r.get('case_start')) else '',
                    ', '.join(prior_q_labels),
                    'Duplicate PID — resolve in Kintone. Already counted in prior quarter(s).'], 1):
                cl = ws_dup.cell(row=dr_dup, column=cc, value=v)
                cl.font = Font(name='Arial', size=10, bold=(cc==1))
                cl.fill = PatternFill('solid', start_color=ORG)
                cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc==7))
            dr_dup += 1
    TBR(ws_dup, 3, max(dr_dup-1, 4), 1, 7)
    for cc, w in zip(range(1,8), [12,14,16,16,12,16,52]):
        ws_dup.column_dimensions[get_column_letter(cc)].width = w

    # ── Corrections Master ────────────────────────────────────────
    ws_corr = wb.create_sheet("Corrections — Master")
    title_row(ws_corr, 1, 1, 7, f"ATX Peace {QUARTER} {FY} — Master Corrections Log", bg=DRED)
    for cc, h in enumerate(['#','Issue Type','Source','Coordinator','PID','Name / Detail','Action Required'], 1):
        H(ws_corr, 2, cc, h, bg=DRED, wrap=True)
    ws_corr.row_dimensions[2].height = 28
    for i, corr in enumerate(corrections, 1):
        bg = corr_color(corr[0])
        for cc, v in enumerate([i] + corr, 1):
            cl = ws_corr.cell(row=i+2, column=cc, value=v)
            cl.font = Font(name='Arial', size=10); cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left', vertical='center', wrap_text=(cc in [6,7]))
    TBR(ws_corr, 2, len(corrections)+2, 1, 7)
    for cc, w in zip(range(1,8), [5,34,20,16,12,44,52]):
        ws_corr.column_dimensions[get_column_letter(cc)].width = w
    ws_corr.freeze_panes = 'C3'

    # ── Serialize to bytes ────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"ATX_Peace_{QUARTER}_{FY.replace(' ','_')}_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    stats = {
        'participants': len(q3_elig),
        '5b_pct':       ls_pct,
        'outreach':     len(out_q3),
        'corrections':  len(corrections),
        'rollover':     len(rollover),
        'circles':      len(circ_sessions),
        'incidents':    len(inc_q3),
        'duplicates':   len(duplicate_q3),
    }
    return {'bytes': buf.getvalue(), 'stats': stats, 'fname': fname}
